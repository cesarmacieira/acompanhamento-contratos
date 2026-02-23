import requests
import json
import time
from datetime import datetime
import os
from dotenv import load_dotenv
load_dotenv()
url_comprasnet = os.environ.get('URL_COMPRASNET')


class BuscadorContratos:
    """
    Classe para buscar contratos na API de Dados Abertos do governo federal
    """
    def __init__(self):
        self.base_url = url_comprasnet
        self.headers = {'accept': '*/*', 'User-Agent': 'Python Script - Busca Contratos'}

    def buscar_contratos(self, orgao_codigo, unidade_gestora, data_inicio=None, data_fim=None, tamanho_pagina=500):
        """
        Busca todos os contratos de um órgão e unidade gestora específicos

        Args:
            orgao_codigo: Código do órgão (ex: 12000)
            unidade_gestora: Código da unidade gestora (ex: 090006)
            data_inicio: Data inicial no formato 'YYYY-MM-DD' (padrão: 2015-01-01)
            data_fim: Data final no formato 'YYYY-MM-DD' (padrão: hoje)
            tamanho_pagina: Quantidade de registros por página (padrão: 500)

        Returns:
            Lista com todos os contratos encontrados
        """
        if data_inicio is None:
            data_inicio = "2015-01-01"
        if data_fim is None:
            data_fim = datetime.now().replace(year=datetime.now().year + 5).strftime("%Y-%m-%d")

        todos_contratos = []
        pagina = 1

        while True:
            try:
                params = {
                    'pagina': pagina,
                    'tamanhoPagina': tamanho_pagina,
                    'codigoOrgao': orgao_codigo,
                    'codigoUnidadeGestora': unidade_gestora,
                    'dataVigenciaInicialMin': data_inicio,
                    'dataVigenciaInicialMax': data_fim,
                }

                response = requests.get(self.base_url, params=params, headers=self.headers, timeout=60)

                if response.status_code == 200:
                    dados = response.json()
                    contratos = dados.get('resultado', [])

                    if not contratos:
                        break

                    quantidade = len(contratos)
                    todos_contratos.extend(contratos)

                    if quantidade < tamanho_pagina:
                        break

                    pagina += 1
                    time.sleep(1)

                elif response.status_code == 404:
                    break
                elif response.status_code == 429:
                    time.sleep(60)
                    continue
                else:
                    break

            except requests.exceptions.Timeout:
                time.sleep(5)
                continue
            except requests.exceptions.RequestException:
                break
            except json.JSONDecodeError:
                break

        return todos_contratos

    def buscar_por_ano(self, orgao_codigo, unidade_gestora, ano):
        """
        Busca contratos de um ano específico
        """
        data_inicio = f"{ano}-01-01"
        data_fim = f"{ano}-12-31"
        return self.buscar_contratos(orgao_codigo, unidade_gestora, data_inicio, data_fim)

    def buscar_multiplos_anos(self, orgao_codigo, unidade_gestora, ano_inicio=2015, ano_fim=None):
        """
        Busca contratos de múltiplos anos
        """
        if ano_fim is None:
            ano_fim = datetime.now().year

        todos_contratos = []
        for ano in range(ano_inicio, ano_fim + 1):
            contratos = self.buscar_por_ano(orgao_codigo, unidade_gestora, ano)
            if contratos:
                todos_contratos.extend(contratos)
            time.sleep(2)

        return todos_contratos

    def salvar_excel(self, contratos, nome_arquivo=None):
        """
        Salva os contratos em um arquivo Excel com formatação
        """
        try:
            import pandas as pd
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        if not contratos:
            return None

        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"contratos_{timestamp}.xlsx"

        df = pd.DataFrame(contratos)

        colunas_principais = [
            'numeroContrato', 'nomeRazaoSocialFornecedor', 'valorGlobal',
            'dataVigenciaInicial', 'dataVigenciaFinal', 'objeto',
            'nomeCategoria', 'nomeTipo', 'processo',
        ]
        colunas_ordenadas = [col for col in colunas_principais if col in df.columns]
        outras_colunas = [col for col in df.columns if col not in colunas_principais]
        df = df[colunas_ordenadas + outras_colunas]

        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Contratos', index=False)
            worksheet = writer.sheets['Contratos']

            if 'valorGlobal' in df.columns:
                col_idx = df.columns.get_loc('valorGlobal') + 1
                for row in range(2, len(df) + 2):
                    worksheet.cell(row=row, column=col_idx).number_format = 'R$ #,##0.00'

            for col_name in ['dataVigenciaInicial', 'dataVigenciaFinal', 'dataHoraInclusao']:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=col_idx).number_format = 'DD/MM/YYYY'

            worksheet.freeze_panes = 'A2'

        return nome_arquivo